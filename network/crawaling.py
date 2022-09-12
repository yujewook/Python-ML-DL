# 쓰레드
# 쓰레드를 안쓴 경우
# import time
# def doing():
#     time.sleep(1)
#     print("실행중")
#
# start = time.time()
# for i  in range(10):
#     doing()
# end = time.time()
# print("실행시간 : " ,(end-start))

import time
import threading
# def doing():
#     time.sleep(1)
#     print("스레드 실행중")
# #스레드 꼭 메인을 만들어줘야한다. if __name__ == "__main__":
# if __name__ == "__main__":
#     start = time.time()
#     threads = []
#     for i in range(10):
#         t = threading.Thread(target=doing)
#         t.start() #실행대기 상태
#         threads.append(t)
#     for thread in threads :
#         thread.join()
#     end = time.time()
#     print("실행시간 : " ,(end-start))

from concurrent import futures #start()join 자동처리
from multiprocessing import pool
# def doing():
#     time.sleep(1)
#     return "스레드 실행중"
# if __name__ == "__main__":
#     start = time.time()
#     results = []
#     with futures.ThreadPoolExecutor() as excutor :
#         for i in range(10):
#             result = excutor.submit(doing)
#             results.append(result)
#     for f in futures.as_completed(results):
#         print(f.result())
#     end = time.time()
#     print("실행시간:" ,(end-start))

# def calc_sum(list):
#     sum = 0
#     for i in range(list[0],list[1]+1):
#         sum += i
#     return sum
# if __name__ == "__main__" :
#     start = time.perf_counter()
#     result = calc_sum([1,10000000])
#     print(result)
#     end = time.perf_counter()
#     print("실행시간 :",(end-start))
    
# def calc_sum(list):
#     sum = 0
#     for i in range(list[0],list[1]+1):
#         sum+=i
#     return sum
# if __name__ == "__main__" :
#     start = time.time()
#     with futures.ThreadPoolExecutor() as excutor :
#         sub = [[1,100000000//2],[100000000//2+1,100000000]] #쓰레드 이등분 한것
#         results = excutor.map(calc_sum,sub)
#     print( sum(results))
#     end = time.time()
#     print("실행시간 :" ,(end-start))
#

from multiprocessing import Pool
#멀티 프로세스
# def calc_sum(list):
#     sum = 0
#     for i in range(list[0],list[1]+1):
#         sum +=1
#     print(sum)
# if __name__  == "__main__":
#     start = time.time()
#     sub = [[1,100000000//2],[100000000//2+1,100000000]]
#     pool = Pool(processes=2)
#     pool.map(calc_sum,sub)
#     pool.close()
#     pool.join()
#     end = time.time()
#     print("실행시간:" ,(end-start))

#쓰레드 동기화 동시에 동작하면 서 공유하는 변수가 있어야한다.
#공유변수
# number = 0
# #락을 만들어야한다.
# lock = threading.Lock() #락객체를 만듬
# def thread1(num):
#     global number
#     lock.acquire()         #락을 걸어버림 -> 다른 스레드 접근 금지
#     for i in range(num+1):
#         number += 1
#     lock.release()         #락을 풀어버림
# def thread2(num):
#     global number
#     lock.acquire()
#     for i in range(num+1):
#         number += 1
#     lock.release()
# #메인
# if __name__ == "__main__":
#     threads = []
#     start = time.time()
#     t1 = threading.Thread(target=thread1 , args=(50000000,))
#     t1.start()
#     threads.append(t1)
#
#     t2 = threading.Thread(target=thread2 , args=(50000000,))
#     t2.start()
#     threads.append(t2)
#     # 조인
#     for thread in threads :
#         thread.join()
#     print(number)
#     end=time.time()
#     print("실행시간 :" ,(end-start) )

# #멀티프로세스 동기화
#     # id 스레드번호 , number 최대값 ,shm 공유메모리, arr numpy.arr, sem 세마포어
# from multiprocessing import shared_memory,Semaphore ,Process#shm 사용하기위한 임포트
# import numpy as np #np 사용하기위해 임포트 
# def calc_sum(id, number , shm , arr, sem ):
#     sum = 0
#     for i in range(number):
#         sum += 1
#     sem.acquire() #임계영역은 하나만 들어온다 -> 다른거 못들어 온다 but 세마포어 자리를획득 -> 여러개 들어온다 -> 충돌이 안나온다
#
#     new_shm= shared_memory.SharedMemory(name=shm)
#                         #리스트 수    타입            버퍼임시 저장소
#     tmp_arr = np.ndarray(arr.shape,dtype=arr.dtype, buffer=new_shm.buf) 
#     tmp_arr[0] += sum
#     sem.release() #세미포어 해제
# if __name__=="__main__":
#     start = time.time()
#     arr = np.array([0])
#     shm = shared_memory.SharedMemory(create=True,size=arr.nbytes)
#     np_shm = np.ndarray(arr.shape, dtype=arr.dtype , buffer=shm.buf)
#     sem =Semaphore()
#     p1 = Process(target=calc_sum, args=(1,500000, shm.name, np_shm, sem))
#     p2 = Process(target=calc_sum, args=(1,500000, shm.name, np_shm, sem))
#     p1.start()
#     p2.start()
#     p1.join()
#     p2.join()
#     end = time.time()
#     print(np_shm[0])
#     print("실행시간",(end-start))
#     shm.close()
#     shm.unlink()
       
       
#네트워크
import urllib.request as req
from urllib.error import URLError
from numpy.testing._private.parameterized import param
from numpy.core.defchararray import title
import openpyxl
from _ast import Lambda
# f=req.urlopen("http://www.daum.net")
# print(f.read(100).decode("utf-8"))

# re = req.Request("http://www.daum.net")
# try :
#     req.urlopen(re)
# except URLError as e :
#     print(e.reason)

#응답을 받늗다
response = req.urlopen("http://www.daum.net")#객체가 넘어온다.
#print(response)
print("url :" , response.geturl())
headers = response.info() # 헤더는 리스트롤 받는다
print("date :" ,headers["date"]) #수정한 날짜
data=response.read()
print(len(data))

#크롤링
#requests
import urllib.request
url = "https://imgnews.pstatic.net/image/109/2022/08/19/0004681395_001_20220819111505656.png?type=w647"

#저장할 이름 확장자를 맞추자
# savename = "baseball.png"
# urllib.request.urlretrieve(url,savename)
# print("저장했습니다")

#파일 저장하는 것 
#url = "http://www.google.com/robots.txt"
#연결하기
#txt = urllib.request.urlopen(url)
#data = txt.read().decode("utf-8")
#print(data)
#파일 출력하기
#with open("robots.txt","w", encoding="utf-8") as f :
#    f.write(data)
#print("저장했습니다")

#RSS
#데이터 갖고 오는 애
import urllib.request as req 
#url parse를 위해서
import urllib.parse as pa
url = "http://www.kma.go.kr/weather/forecast/mid-term-rss3.jsp"
#딕셔러리로 만든다 변수를 만들고 
values ={
    "stnId" : 109
    
    }
# params = pa.urlencode(values)
# url = url +"?"+params #url 연결해주면 서 ? 받는 get 방식?
# #print(url)
# #객체를 받기 때문에
# data = req.urlopen(url).read().decode("utf-8")
# #print(data)
#
# #beautifulSoup
# from bs4 import BeautifulSoup as bs
# html = """
#      <html>
#         <head>
#             <meta charset="utf-8">
#             <title>HTML TEST</title>
#         </head>
#         <body>
#             <h2> html 연습 </h2>
#             <p id ="first"> 짜장면</p>
#             <p class ="strong"> 짬뽕</p>
#             <p class ="point Strong"> 탕수육</p>
#             <p id="second"> 울면</p>
#         </body>
#     </html>
# """
# soup = bs( html,"html.parser")
# #soup으로 찾아라 .string 글자를 찾아라
# print("<<"+ soup.find("h2").string+ ">>")
# print("<<"+soup.html.body.h2.string+">>")
#
# h2 = soup.find("h2")
# body = h2.parent
# #print(body)
# html = body.parent
# print(html)
# #h2 다음 것은 <p>
# p1 = h2.next_sibling.next_sibling
# print(p1.string)
# nodes = body.childern #자식 노드들을 갖고안다
# # for node in nodes :
# #     print(node.string)
# ps = soup.find_all("p")
# for p in ps : 
#     print(p.string) 
# p1 = soup.find(id="first")
# print(p1.string)
# p2 = soup.find(id="second")
# print(p2.string)
#             ##html에서 class 선택자 
# pp = soup.find_all(class_="strong")
# for p in pp:
#     print(p.string)
#
# #select()
# #select_one()하나만 찾을때
# html ="""
# <html>
#     <head>
#         <meta charset="utf-8">
#         <title> html 연습 </title>
#     </head>
#     <body>
#         <h2 class="title">웹 스크랩핑 </h2>
#         <p id="name">html</p>
#         <p class="subject"><a>xml</a></p>
#         <p class="subject"><b><a>json</a></b></p>
#         <a>CDATA</a>
#     </body>
# </html>
# """
#
# soup =bs(html,"html.parser")
# h2 = soup.select_one("body h2")
# print("<<" +h2.string +">>")
#
# ps = soup.select(".subject")
# for p in ps:
#     print(p.string)
#
# pp = soup.select("body p")
# for p in pp:
#     print(p.string)
#
# ps = soup.select("p a") #자손 선택자
# for p in ps :
#     print(p.string)
# p = soup.select("p>a") #자식 선택자  
# for p in pp :
#     print (p.string) #xml하나만 선택
#
# p=soup.select_one("p#name")
# print(p.string)
#
# p=soup.select_one("p.subject")
# print(p.string)
#
# data = req.urlopen(url).read().decode("utf-8")
# #print(data)
#
# #기상 데이터 파일화 하기
# with open("whether.txt","w",encoding="utf-8") as f:
#     f.write(data)
#
# #데이터 짤라보기
# soup = bs(data,"html.parser")
# title = soup.find("title")
# #기상청 데이터 title태그 꺼 갖고옴
# print("<<" + title.string +">>")
#
# cities = soup.find_all("city")
# # for city in cities :
# #     print(city.string)
#
# datas = soup.find_all("data")
# for data in datas:
#     print (data.parent.city.string ,end="\t")
#     #xml에서 tmef태그로 (시간) 되있는 데이터 갖고 오기  , 날씨데이터
#     print(data.tmef.string   , "\t" ,data.wf.string,
#           # 최저 기온  
#           "\t" , data.tmx.string)
#
# print()

# #json 데이터 중괄호로 묶여있다. -> 딕셔너리랑 비슷하다.
# import json
# import os.path
# url = "http://api.github.com/repositories"
# filename = "github.txt" 
#
# if not os.path.exists(filename):
#     req.urlretrieve(url, filename)
#     # 파일 읽어오기
# items = json.load(open(filename , "rt",encoding="utf-8"))
#
# # json { {,,,} } 되어 있으니까 반복문을 돌리자
# for item in items :
#     #item 안에 자료 id,name key 꺼내기
#                                     #json 안에 json데이터 꺼내기
#     print(item["id"],item["name"],item["owner"]["login"])


# csv데이터
import csv, codecs
filename = "test.csv"
#파일연결해서 csv파일 쓰기 모드
# write = codecs.open(filename,"w",encoding="utf-8")
# # csv 에쓰는 것 구분자 만들기 delimeter=","
# writer = csv.writer(write, delimiter=",")
# #쓰기 
# writer.writerow(["아이디","이름","가격"])
# writer.writerow(["1000","HDD",200000]) 
# writer.writerow(["1001","SDD",300000]) 
# writer.writerow(["1002","Monitor",150000]) 
# writer.writerow(["1003","Mous",20000]) 
# writer.writerow(["1004","keyboard",300000]) 
# print("생성완료")
#
# #읽을려면 한번 닫아 줘야한다.
# write.close()
# # 다시 연결 하고 읽기 모드
# readcsv= codecs.open(filename , "r" ,encoding="utf-8").read()
# #데이터 리스트로 만들기 변수 선언하고
# data = []
# # 한줄씩 짜르기
# rows = readcsv.split("\r\n")
#
#
# #print(rows) 자료 확인하고
# for row in rows:
#     if row == "":
#         break
#     cells =row.split(",")
#     data.append(cells) #data라는 변수에 넣기
#
# print(data)

#한줄 씩 만들기
# for d in data : # 리스트 형식
#     print(d[0],"\t",d[1],"\t",d[2])
    

#엑셀 데이터
import openpyxl
#엑셀 xlsx 확장자 일때 가능하다.
filename = "stat_100701.xlsx"
# work book을 만들고
wb = openpyxl.load_workbook(filename)
#읽는 방법
#1 ws = wb.worksheets[0]
#2
ws = wb.active
# print(ws.rows) 객체를 보여준다

# 이중 포문으로 뽑아준다. 출력용
# for row in ws.rows : # 리스트로 받아준다
#     for data in row : 
#         if data.value == None:
#             print("",end="\t")
#         else:     
#             print(data.value , end="\t")
#     print()
    
#다시 가공해서 엑셀에 넣어주기
data = []
            # 워크쉬트에 rows
for row in ws.rows :
    if row[9] != None and row[10] != None :
        data.append( [row[9].value , row[10].value] )
del(data[0:4]) # none 나오는 부분을 지우기 위해서 지움

#sort 해주기 위해서
#        return 값이 있고 무엇을 소트할지 함수를 만들어준다 (람다로 )
                        # 인구 밀도를 기준으로 sort
                                        # 역으로 
data = sorted(data , key = lambda x:x[1] , reverse = True)


#가공 완료
for d in data : 
    print(d)




#가공한 데이터 저장할 파일 만들어 주기
savefile = "population.xlsx"
swb = openpyxl.Workbook()
#활성화
sws = swb.active
#sws = swb.create_sheet(title="인구") 주석처리 되면 원래 시트에 들어 가게 된다
# 인덱스 붙여넣기 어디에다 넣을지 위치를 만들어 줘야 하기 떄문에 인덱스가 필요하다
for i,d in enumerate(data) :
    sws.cell(row=i+1,column=1,value=d[0])
    sws.cell(row=i+1,column=2,value=d[1])    
swb.save(savefile)
        
        
        
        
        


















 










    

    
    